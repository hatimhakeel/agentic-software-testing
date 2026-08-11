#!/usr/bin/env python3
"""Highlight updated cells between an old workbook and a single-sheet new workbook.

Only changed or added cells that exist in the new workbook are highlighted
FFF2CC. Deleted rows/cells from the old worksheet are never inserted, copied,
restored, represented, or highlighted in the output workbook.
"""

from __future__ import annotations

import argparse
import copy
import json
import re
import sys
from dataclasses import dataclass
from difflib import SequenceMatcher
from pathlib import Path
from typing import Any, Iterable

from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.worksheet.worksheet import Worksheet

CHANGED_FILL = PatternFill(fill_type="solid", fgColor="FFF2CC")
MIN_FUZZY_SCORE = 0.74


@dataclass(frozen=True)
class SheetMatch:
    old_sheet_name: str
    new_sheet_name: str
    method: str
    score: float


@dataclass
class DiffStats:
    changed_or_added_cells: int = 0


def normalize_sheet_name(name: str) -> str:
    text = name.lower().strip()
    text = re.sub(r"[\s_\-\.\,;:]+", " ", text)
    text = re.sub(r"[^a-z0-9 ]+", "", text)
    return re.sub(r"\s+", " ", text).strip()


def ratio(a: str, b: str) -> float:
    return SequenceMatcher(None, a, b).ratio()


def find_matching_sheet(old_names: list[str], new_name: str) -> SheetMatch:
    if new_name in old_names:
        return SheetMatch(new_name, new_name, "exact", 1.0)

    lower_matches = [name for name in old_names if name.lower() == new_name.lower()]
    if len(lower_matches) == 1:
        return SheetMatch(lower_matches[0], new_name, "case-insensitive", 0.99)
    if len(lower_matches) > 1:
        raise ValueError(f"multiple case-insensitive sheet matches for {new_name!r}: {lower_matches}")

    new_norm = normalize_sheet_name(new_name)
    norm_matches = [name for name in old_names if normalize_sheet_name(name) == new_norm]
    if len(norm_matches) == 1:
        return SheetMatch(norm_matches[0], new_name, "normalized", 0.98)
    if len(norm_matches) > 1:
        raise ValueError(f"multiple normalized sheet matches for {new_name!r}: {norm_matches}")

    partial_candidates: list[tuple[float, str]] = []
    for old_name in old_names:
        old_norm = normalize_sheet_name(old_name)
        if not old_norm or not new_norm:
            continue
        if new_norm in old_norm or old_norm in new_norm:
            containment = min(len(new_norm), len(old_norm)) / max(len(new_norm), len(old_norm))
            score = max(ratio(new_norm, old_norm), containment)
            partial_candidates.append((score, old_name))
    if partial_candidates:
        partial_candidates.sort(key=lambda item: (-item[0], item[1].lower()))
        best_score = partial_candidates[0][0]
        tied = [name for score, name in partial_candidates if abs(score - best_score) < 1e-9]
        if len(tied) > 1:
            raise ValueError(f"multiple equally strong partial sheet matches for {new_name!r}: {tied}")
        return SheetMatch(partial_candidates[0][1], new_name, "partial", best_score)

    fuzzy_candidates = [(ratio(new_norm, normalize_sheet_name(old_name)), old_name) for old_name in old_names]
    fuzzy_candidates.sort(key=lambda item: (-item[0], item[1].lower()))
    best_score, best_name = fuzzy_candidates[0]
    if best_score < MIN_FUZZY_SCORE:
        raise ValueError(
            f"no confident sheet match found for {new_name!r}; available old workbook sheets: {old_names}"
        )
    tied = [name for score, name in fuzzy_candidates if abs(score - best_score) < 1e-9]
    if len(tied) > 1:
        raise ValueError(f"multiple equally strong fuzzy sheet matches for {new_name!r}: {tied}")
    return SheetMatch(best_name, new_name, "fuzzy", best_score)


def is_blank(value: Any) -> bool:
    return value is None or value == ""


def comparable_value(value: Any) -> str:
    if is_blank(value):
        return ""
    if hasattr(value, "isoformat"):
        try:
            return value.isoformat()
        except Exception:
            pass
    return str(value).strip()


def is_numbering_like(value: Any) -> bool:
    """Detect serial/index-style values to ignore during row alignment.

    Intentionally broad: plain integers, decimals, Roman numerals, one-letter
    bullets, and common enumerations such as 1., (a), A-1, or 001.
    """
    if is_blank(value):
        return False
    if isinstance(value, bool):
        return False
    if isinstance(value, int):
        return True
    if isinstance(value, float):
        return value.is_integer()
    text = str(value).strip()
    if not text:
        return False
    patterns = [
        r"^\(?\d+\)?[\.)\-:]?$",
        r"^\d+[\.\-]\d+([\.\-]\d+)*$",
        r"^[A-Za-z]\)?[\.)\-:]$",
        r"^\(?[ivxlcdmIVXLCDM]+\)?[\.)\-:]?$",
        r"^(sr\.?|serial|s/?n|no\.?)\s*\d+$",
    ]
    return any(re.match(pattern, text) for pattern in patterns)


def used_bounds(ws: Worksheet) -> tuple[int, int, int, int]:
    max_row = ws.max_row or 1
    max_col = ws.max_column or 1
    min_row = 1
    min_col = 1

    last_row = 1
    last_col = 1
    found = False
    for row in ws.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col):
        for cell in row:
            if not is_blank(cell.value):
                found = True
                last_row = max(last_row, cell.row)
                last_col = max(last_col, cell.column)
    if not found:
        return 1, 1, 1, 1
    return min_row, last_row, min_col, last_col


def row_values(ws: Worksheet, row_idx: int, max_col: int) -> list[Any]:
    return [ws.cell(row=row_idx, column=col_idx).value for col_idx in range(1, max_col + 1)]


def is_numbering_column(ws: Worksheet, col_idx: int, max_row: int) -> bool:
    header = comparable_value(ws.cell(row=1, column=col_idx).value).lower()
    header = re.sub(r"[^a-z0-9]+", " ", header).strip()
    header_terms = {"no", "number", "serial", "sr", "s n", "sn", "index", "row", "item no", "sl no"}
    if header in header_terms:
        return True

    ints: list[int] = []
    for row_idx in range(2, max_row + 1):
        value = ws.cell(row=row_idx, column=col_idx).value
        if isinstance(value, bool) or is_blank(value):
            continue
        if isinstance(value, int):
            ints.append(value)
        elif isinstance(value, float) and value.is_integer():
            ints.append(int(value))
        elif isinstance(value, str) and re.match(r"^\d+$", value.strip()):
            ints.append(int(value.strip()))
    if len(ints) < 3:
        return False
    diffs = [b - a for a, b in zip(ints, ints[1:])]
    sequential_diffs = sum(1 for diff in diffs if diff == 1)
    return bool(diffs) and sequential_diffs / len(diffs) >= 0.8


def detect_numbering_columns(old_ws: Worksheet, new_ws: Worksheet, max_col: int) -> set[int]:
    _, old_max_row, _, _ = used_bounds(old_ws)
    _, new_max_row, _, _ = used_bounds(new_ws)
    return {
        col_idx
        for col_idx in range(1, max_col + 1)
        if is_numbering_column(old_ws, col_idx, old_max_row) or is_numbering_column(new_ws, col_idx, new_max_row)
    }


def row_signature(values: Iterable[Any], numbering_cols: set[int]) -> str:
    tokens: list[str] = []
    for col_idx, value in enumerate(values, start=1):
        if col_idx in numbering_cols:
            continue
        text = comparable_value(value).lower()
        # Do not use formulas for row alignment. Formula references commonly shift
        # after row deletions and can create false row matches. Formulas are still
        # compared cell-by-cell after rows are aligned.
        if text.startswith("="):
            continue
        if text:
            tokens.append(text)
    return " | ".join(tokens)


def row_similarity(old_sig: str, new_sig: str) -> float:
    """Score row similarity while tolerating deleted cells inside a row."""
    if old_sig == new_sig:
        return 1.0
    if not old_sig and not new_sig:
        return 1.0
    if not old_sig or not new_sig:
        return 0.0
    old_tokens = [token.strip() for token in old_sig.split("|") if token.strip()]
    new_tokens = [token.strip() for token in new_sig.split("|") if token.strip()]
    old_set = set(old_tokens)
    new_set = set(new_tokens)
    if old_set and new_set:
        overlap = len(old_set & new_set)
        coverage = overlap / min(len(old_set), len(new_set))
        jaccard = overlap / len(old_set | new_set)
    else:
        coverage = 0.0
        jaccard = 0.0
    return max(ratio(old_sig, new_sig), coverage, jaccard)


def row_has_content(ws: Worksheet, row_idx: int, max_col: int) -> bool:
    return any(not is_blank(ws.cell(row=row_idx, column=col_idx).value) for col_idx in range(1, max_col + 1))


def fill_cell(cell, fill: PatternFill) -> None:
    cell.fill = copy.copy(fill)


def build_row_alignment(old_ws: Worksheet, new_ws: Worksheet, max_col: int, numbering_cols: set[int]) -> list[tuple[str, int | None, int | None]]:
    _, old_max_row, _, _ = used_bounds(old_ws)
    _, new_max_row, _, _ = used_bounds(new_ws)
    old_sigs = [row_signature(row_values(old_ws, row_idx, max_col), numbering_cols) for row_idx in range(1, old_max_row + 1)]
    new_sigs = [row_signature(row_values(new_ws, row_idx, max_col), numbering_cols) for row_idx in range(1, new_max_row + 1)]

    matcher = SequenceMatcher(None, old_sigs, new_sigs, autojunk=False)
    alignment: list[tuple[str, int | None, int | None]] = []

    for tag, i1, i2, j1, j2 in matcher.get_opcodes():
        if tag == "equal":
            for offset in range(i2 - i1):
                alignment.append(("match", i1 + offset + 1, j1 + offset + 1))
        elif tag == "delete":
            for old_idx in range(i1 + 1, i2 + 1):
                alignment.append(("deleted_row", old_idx, None))
        elif tag == "insert":
            for new_idx in range(j1 + 1, j2 + 1):
                alignment.append(("added_row", None, new_idx))
        elif tag == "replace":
            # Pair rows in order so deleted-row placeholders do not scramble the
            # logical worksheet sequence. A row can still match when one or more
            # non-numbering cells were deleted from it.
            matched_new: set[int] = set()
            next_new_start = j1
            pending_added: list[int] = []

            for oi in range(i1, i2):
                best_score = 0.0
                best_nj: int | None = None
                for nj in range(next_new_start, j2):
                    if nj in matched_new:
                        continue
                    score = row_similarity(old_sigs[oi], new_sigs[nj])
                    # Prefer rows sharing at least one non-numbering, non-formula token.
                    old_tokens = {token.strip() for token in old_sigs[oi].split("|") if token.strip()}
                    new_tokens = {token.strip() for token in new_sigs[nj].split("|") if token.strip()}
                    if old_tokens and new_tokens and not (old_tokens & new_tokens):
                        score = min(score, 0.25)
                    if score > best_score:
                        best_score = score
                        best_nj = nj
                if best_nj is not None and best_score >= 0.55:
                    for skipped_nj in range(next_new_start, best_nj):
                        if skipped_nj not in matched_new:
                            pending_added.append(skipped_nj)
                            matched_new.add(skipped_nj)
                    alignment.append(("match", oi + 1, best_nj + 1))
                    matched_new.add(best_nj)
                    next_new_start = best_nj + 1
                else:
                    alignment.append(("deleted_row", oi + 1, None))

            for nj in range(j1, j2):
                if nj not in matched_new:
                    pending_added.append(nj)
                    matched_new.add(nj)
            for nj in sorted(set(pending_added)):
                alignment.append(("added_row", None, nj + 1))
        else:
            raise RuntimeError(f"unexpected SequenceMatcher tag: {tag}")
    return alignment


def compare_values(old_value: Any, new_value: Any) -> bool:
    return comparable_value(old_value) == comparable_value(new_value)


def highlight_workbooks(old_path: Path, new_path: Path, output_path: Path) -> dict[str, Any]:
    try:
        old_wb = load_workbook(old_path, data_only=False)
    except Exception as exc:
        raise RuntimeError(f"could not open old workbook {old_path}: {exc}") from exc
    try:
        new_wb = load_workbook(new_path, data_only=False)
    except Exception as exc:
        raise RuntimeError(f"could not open new workbook {new_path}: {exc}") from exc

    if len(new_wb.sheetnames) != 1:
        raise ValueError(
            f"new workbook must contain exactly one worksheet; found {len(new_wb.sheetnames)}: {new_wb.sheetnames}"
        )

    new_sheet_name = new_wb.sheetnames[0]
    match = find_matching_sheet(old_wb.sheetnames, new_sheet_name)
    old_ws = old_wb[match.old_sheet_name]
    new_ws = new_wb[new_sheet_name]

    _, old_max_row, _, old_max_col = used_bounds(old_ws)
    _, new_max_row, _, new_max_col = used_bounds(new_ws)
    max_col = max(old_max_col, new_max_col)
    numbering_cols = detect_numbering_columns(old_ws, new_ws, max_col)

    stats = DiffStats()
    alignment = build_row_alignment(old_ws, new_ws, max_col, numbering_cols)

    # Deleted rows are intentionally ignored in the output. Use the alignment only
    # to compare rows that exist in the new workbook or to highlight added rows.
    processed_new_rows: set[int] = set()
    for kind, old_row, new_row in alignment:
        if new_row is None:
            continue
        if new_row in processed_new_rows:
            continue
        processed_new_rows.add(new_row)
        if kind == "added_row" or old_row is None:
            for col_idx in range(1, max_col + 1):
                if col_idx in numbering_cols and is_numbering_like(new_ws.cell(row=new_row, column=col_idx).value):
                    continue
                cell = new_ws.cell(row=new_row, column=col_idx)
                if not is_blank(cell.value):
                    fill_cell(cell, CHANGED_FILL)
                    stats.changed_or_added_cells += 1
            continue

        for col_idx in range(1, max_col + 1):
            old_cell = old_ws.cell(row=old_row, column=col_idx)
            new_cell = new_ws.cell(row=new_row, column=col_idx)
            old_value = old_cell.value
            new_value = new_cell.value

            if col_idx in numbering_cols and is_numbering_like(old_value) and is_numbering_like(new_value):
                continue
            if is_blank(old_value) and is_blank(new_value):
                continue
            if not is_blank(old_value) and is_blank(new_value):
                # Deleted cells must not be restored, copied, or highlighted.
                continue
            if is_blank(old_value) and not is_blank(new_value):
                fill_cell(new_cell, CHANGED_FILL)
                stats.changed_or_added_cells += 1
                continue
            if not compare_values(old_value, new_value):
                fill_cell(new_cell, CHANGED_FILL)
                stats.changed_or_added_cells += 1

    output_path.parent.mkdir(parents=True, exist_ok=True)
    new_wb.save(output_path)

    return {
        "output": str(output_path),
        "new_sheet_name": match.new_sheet_name,
        "matched_old_sheet_name": match.old_sheet_name,
        "sheet_matching_method": match.method,
        "sheet_matching_score": round(match.score, 4),
        "changed_or_added_cells_highlighted_fff2cc": stats.changed_or_added_cells,
        "deleted_cells_and_rows_included_in_output": False,
        "deleted_content_policy": "deleted rows/cells are not inserted, restored, copied, or highlighted",
    }


def parse_args(argv: list[str]) -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Highlight Excel differences in a copy of the new workbook.")
    parser.add_argument("--old", required=True, help="Path to old/baseline/source .xlsx workbook")
    parser.add_argument("--new", required=True, help="Path to new/updated/output .xlsx workbook")
    parser.add_argument("--output", required=True, help="Path for highlighted output .xlsx workbook")
    return parser.parse_args(argv)


def main(argv: list[str]) -> int:
    args = parse_args(argv)
    try:
        summary = highlight_workbooks(Path(args.old), Path(args.new), Path(args.output))
    except Exception as exc:
        print(f"ERROR: {exc}", file=sys.stderr)
        return 1
    print(json.dumps(summary, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main(sys.argv[1:]))